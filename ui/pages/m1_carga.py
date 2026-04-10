"""
ui/pages/m1_carga.py
====================
Página de carga de datos para el Modelo M1.
Permite seleccionar el archivo Excel y previsualizar los registros.
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
from ui.theme import COLORS, FONTS, DIMS
from core.models.m1_absoluto import procesar_m1, calcular_resumen_metricas

class M1CargaPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS.bg_main)
        self.app = master
        self.df_base = None
        self.df_monitoreo = None
        self.meta = {}

        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        # Header
        header = ctk.CTkFrame(self, fg_color=COLORS.bg_card, corner_radius=0, height=DIMS.topbar_height)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)

        ctk.CTkButton(
            header, text="← Configuración", font=(FONTS.family, FONTS.size_sm),
            fg_color="transparent", text_color=COLORS.primary, hover_color=COLORS.bg_main,
            width=120, height=32, command=lambda: self.app.navegar("m1_config")
        ).place(x=10, y=14)

        ctk.CTkLabel(
            header, text="Modelo M1: Carga de Datos",
            font=(FONTS.family, FONTS.size_md, "bold"), text_color=COLORS.primary
        ).place(relx=0.5, rely=0.5, anchor="center")

        # Cuerpo
        self.cuerpo = ctk.CTkFrame(self, fg_color="transparent")
        self.cuerpo.grid(row=1, column=0, sticky="nsew", padx=40, pady=20)
        self.cuerpo.grid_columnconfigure(0, weight=1)

        # Zona de Selección
        self.card_selección = ctk.CTkFrame(self.cuerpo, fg_color=COLORS.bg_card, corner_radius=DIMS.card_radius, border_width=1, border_color=COLORS.border)
        self.card_selección.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        
        ctk.CTkLabel(
            self.card_selección, text="Selecciona el archivo Excel M1 con tus datos",
            font=(FONTS.family, FONTS.size_sm), text_color=COLORS.text_secondary
        ).pack(pady=(30, 10))

        self.btn_seleccionar = ctk.CTkButton(
            self.card_selección, text="📂 Seleccionar Archivo",
            font=(FONTS.family, FONTS.size_sm, "bold"), fg_color=COLORS.primary,
            text_color="white", height=40, command=self._seleccionar_archivo
        )
        self.btn_seleccionar.pack(pady=(0, 30))
        
        self.path_label = ctk.CTkLabel(self.card_selección, text="Ningún archivo seleccionado", font=(FONTS.family, FONTS.size_xs))
        self.path_label.pack(pady=(0, 10))

        # Zona Resumen (se llena al cargar)
        self.zona_resumen = ctk.CTkFrame(self.cuerpo, fg_color="transparent")
        self.zona_resumen.grid_columnconfigure((0, 1), weight=1)

    def _seleccionar_archivo(self):
        from core.io_excel import leer_excel_m1
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.path_label.configure(text=os.path.basename(path))
            self.current_path = path

            # 1. Lectura de metadatos desde el Excel (Sincronización Total)
            import openpyxl
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                if "Modelo_LBEn" in wb.sheetnames:
                    ws = wb["Modelo_LBEn"]
                    # Leemos de las celdas D5-D7 y M5-M6 donde grabamos los resultados
                    metadata = {
                        "nombre": ws["D5"].value or self.app.session.get("nombre", ""),
                        "fuente": ws["D6"].value or self.app.session.get("fuente", ""),
                        "unidad": ws["D7"].value or self.app.session.get("unidad", ""),
                        "pb_ini": ws["M5"].value or self.app.session.get("pb_ini", ""),
                        "pb_fin": ws["M6"].value or self.app.session.get("pb_fin", ""),
                    }
                    self.app.session.update(metadata)
            except: pass

            # 2. Conteo de registros REALES (Solo celdas con números)
            df_base, df_mon, meta_f, errs = leer_excel_m1(path)
            
            self.df_base = df_base
            self.df_monitoreo = df_mon
            
            n_base = 0
            if df_base is not None and not df_base.empty:
                # Contamos filas donde la columna de consumo (index 1) no esté vacía ni sea texto
                n_base = pd.to_numeric(df_base.iloc[:, 1], errors='coerce').notna().sum()

            n_mon = 0
            if df_mon is not None and not df_mon.empty:
                # Lo mismo para monitoreo
                n_mon = pd.to_numeric(df_mon.iloc[:, 1], errors='coerce').notna().sum()

            self.app.session["n_pb"] = n_base
            self.app.session["n_pr"] = n_mon
            self.app.session["n_cols"] = len(df_base.columns) if df_base is not None else 0
            
            self._mostrar_resumen()

    def _mostrar_resumen(self):
        # Limpiar zona resumen
        for widget in self.zona_resumen.winfo_children(): widget.destroy()
        self.zona_resumen.grid(row=1, column=0, sticky="nsew")

        # Card Info Proyecto (Izquierda)
        card_info = ctk.CTkFrame(self.zona_resumen, fg_color=COLORS.bg_card, corner_radius=DIMS.card_radius, border_width=1, border_color=COLORS.border)
        card_info.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        
        ctk.CTkLabel(card_info, text="Información del Proyecto", font=(FONTS.family, FONTS.size_sm, "bold"), text_color=COLORS.primary).pack(pady=10)
        
        # Sincronizar con la sesión para que NO salgan los guiones
        entidad = self.meta.get('entidad') or self.app.session.get('nombre', '---')
        items = [
            (f"🏢 Entidad: {self.app.session.get('nombre', 'm1p10')}",),
            (f"⚡ Fuente: {self.app.session.get('fuente', 'electricidad')}",),
            (f"📐 Unidad: {self.app.session.get('unidad', 'kWh')}",),
            (f"📅 Periodo Base: {self.app.session.get('pb_ini', '01/2026')} - {self.app.session.get('pb_fin', '12/2026')}",)
        ]
        for i, (txt,) in enumerate(items):
            pady = 2 if i < len(items)-1 else (2, 20)
            ctk.CTkLabel(card_info, text=txt, font=(FONTS.family, FONTS.size_xs), text_color=COLORS.text_primary).pack(anchor="w", padx=40, pady=pady)

        # Card Resumen de Datos (Derecha)
        card_stats = ctk.CTkFrame(self.zona_resumen, fg_color=COLORS.bg_card, corner_radius=DIMS.card_radius, border_width=1, border_color=COLORS.border)
        card_stats.grid(row=0, column=1, sticky="nsew", padx=(12, 0))
        
        ctk.CTkLabel(card_stats, text="Resumen de Datos", font=(FONTS.family, FONTS.size_sm, "bold"), text_color=COLORS.primary).pack(pady=10)
        
        stats = [
            (f"📊 Registros Periodo Base: {self.app.session.get('n_pb', 0)}",),
            (f"📈 Registros Monitoreo: {self.app.session.get('n_pr', 0)}",),
            (f"✅ Columnas encontradas: {self.app.session.get('n_cols', 0)}",)
        ]
        for i, (txt,) in enumerate(stats):
            pady = 2 if i < len(stats)-1 else (2, 20)
            ctk.CTkLabel(card_stats, text=txt, font=(FONTS.family, FONTS.size_xs), text_color=COLORS.text_primary).pack(anchor="w", padx=40, pady=pady)

        # Botón Procesar
        ctk.CTkButton(
            self.cuerpo, text="⚙️ Calcular Línea Base y Desempeño",
            font=(FONTS.family, FONTS.size_md, "bold"), fg_color=COLORS.accent,
            text_color=COLORS.primary, height=48, command=self._procesar_m1
        ).grid(row=2, column=0, pady=32)

    def _procesar_m1(self):
        if self.df_base is None:
            messagebox.showwarning("Sin datos", "Por favor carga un archivo Excel primero.")
            return

        try:
            # Ejecutar el motor de cálculo
            df_lben, df_mon_res, df_base_final, df_excluidos = procesar_m1(self.df_base, self.df_monitoreo)
            
            # Calcular métricas para la ficha técnica
            metricas = calcular_resumen_metricas(df_lben, self.df_base)
            
            # Guardar todo en sesión
            self.app.session["df_base_raw"] = self.df_base
            self.app.session["df_base_final"] = df_base_final
            self.app.session["df_excluidos"] = df_excluidos
            self.app.session["df_lben"] = df_lben
            self.app.session["df_monitoreo"] = df_mon_res
            self.app.session["metricas_m1"] = metricas
            self.app.session["meta_m1"] = self.app.session # Usar la sesión como metadatos
            self.app.session["excel_path"] = self.current_path
            
            # Navegar a resultados
            self.app.navegar("m1_resultados")
            
        except Exception as e:
            messagebox.showerror("Error en el cálculo", f"No se pudo procesar el modelo: {e}")
