"""
ui/pages/m2_carga.py
====================
Página de carga de datos para el Modelo M2.
Réplica exacta de la estética M1, con Ficha de Resumen integrada.
"""

import customtkinter as ctk
import os
import pandas as pd
from tkinter import filedialog, messagebox
from ui.theme import COLORS, FONTS, DIMS
from core.models.m2_cociente import procesar_m2, calcular_resumen_metricas_m2
from core.io_excel import leer_excel_m2
import threading

class M2CargaPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLORS.bg_main)
        self.app = master
        self.df_base = None
        self.df_monitoreo = None
        self.current_path = None
        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        # Topbar Corporativa
        topbar = ctk.CTkFrame(self, fg_color=COLORS.bg_card, height=DIMS.topbar_height, corner_radius=0)
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(1, weight=1)
        
        ctk.CTkFrame(topbar, fg_color=COLORS.accent, height=2).place(relx=0, rely=1.0, relwidth=1.0, anchor="sw")

        ctk.CTkButton(topbar, text="← Configuración", font=(FONTS.family, FONTS.size_sm),
                      fg_color="transparent", text_color=COLORS.primary, 
                      command=lambda: self.app.navegar("m2_config")).grid(row=0, column=0, padx=16)
        
        ctk.CTkLabel(topbar, text="Modelo M2: Cociente de Valores — Carga de Datos", font=(FONTS.family, FONTS.size_md, "bold"), 
                     text_color=COLORS.primary).grid(row=0, column=1, sticky="w")

        # Cuerpo con Scroll (idéntico a M1)
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent", corner_radius=0)
        self.scroll.grid(row=1, column=0, sticky="nsew")
        self.scroll.grid_columnconfigure(0, weight=1)

        # Card de Selección
        self.card_sel = ctk.CTkFrame(self.scroll, fg_color=COLORS.bg_card, corner_radius=DIMS.card_radius, border_width=1, border_color=COLORS.border)
        self.card_sel.grid(row=0, column=0, sticky="ew", padx=40, pady=(20, 20))
        
        ctk.CTkLabel(self.card_sel, text="Selecciona el archivo Excel M2 con tus datos diligenciados", 
                     font=(FONTS.family, FONTS.size_sm), text_color=COLORS.text_secondary).pack(pady=(30, 15))
        
        self.btn_sel = ctk.CTkButton(self.card_sel, text="📁 Seleccionar Archivo", font=(FONTS.family, FONTS.size_md, "bold"),
                                     fg_color=COLORS.primary, height=44, command=self._seleccionar_archivo)
        self.btn_sel.pack(pady=(0, 10))
        
        self.lbl_path = ctk.CTkLabel(self.card_sel, text="Ningún archivo seleccionado", font=(FONTS.family, FONTS.size_xs), text_color=COLORS.text_secondary)
        self.lbl_path.pack(pady=(0, 30))

        # Zona para la Ficha de Resumen (aparecerá dinámicamente)
        self.zona_resumen = ctk.CTkFrame(self.scroll, fg_color="transparent")
        self.zona_resumen.grid(row=1, column=0, sticky="nsew")
        self.zona_resumen.grid_columnconfigure(0, weight=1)

    def _seleccionar_archivo(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path: return
        
        self.current_path = path
        self.lbl_path.configure(text=os.path.basename(path), text_color=COLORS.primary)

        # 1. Sincronización Inteligente de Metadatos
        import openpyxl
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            if "Modelo_LBEn" in wb.sheetnames:
                ws = wb["Modelo_LBEn"]
                self.app.session.update({
                    "nombre": ws["D5"].value or self.app.session.get("nombre", ""),
                    "fuente": ws["D6"].value or self.app.session.get("fuente", ""),
                    "unidad": ws["D7"].value or self.app.session.get("unidad", ""),
                    "zona": ws["D8"].value or self.app.session.get("zona", ""),
                    "area": ws["D9"].value or self.app.session.get("area", "No disponible"),
                    "var_relevante_nom": ws["D10"].value or "Indicador",
                    "var_relevante_uni": ws["D11"].value or "X"
                })
        except: pass

        # 2. Lectura y Conteos de la Plantilla M2
        df_b, df_m, _, _ = leer_excel_m2(path)
        self.df_base = df_b
        self.df_monitoreo = df_m

        if df_b is not None and not df_b.empty:
            # En M2 la energía está en la columna 1 (B es index, C es [0], D es [1]) -> No, C es 1.
            # Según io_excel, df_b suele tener Fecha como índice o primera col.
            n_pb = pd.to_numeric(df_b.iloc[:,1], errors='coerce').notna().sum()
            self.app.session["n_pb"] = n_pb
            self.app.session["n_cols"] = len(df_b.columns)
            self.app.session["pb_ini"] = str(df_b["Fecha"].iloc[0])
            self.app.session["pb_fin"] = str(df_b["Fecha"].iloc[-1])
        
        if df_m is not None and not df_m.empty:
            n_pr = pd.to_numeric(df_m.iloc[:,1], errors='coerce').notna().sum()
            self.app.session["n_pr"] = n_pr

        self._mostrar_resumen()

    def _mostrar_resumen(self):
        for w in self.zona_resumen.winfo_children(): w.destroy()
        
        # Ficha de Resumen Premium (Réplica M1)
        card = ctk.CTkFrame(self.zona_resumen, fg_color=COLORS.bg_card, corner_radius=DIMS.card_radius, border_width=1, border_color=COLORS.border)
        card.grid(row=0, column=0, padx=40, sticky="ew")
        card.grid_columnconfigure((0, 1), weight=1)

        # Encabezado: Tipo de Modelo
        ctk.CTkLabel(card, text="M2 (Cociente de Valores Medidos)", font=(FONTS.family, FONTS.size_md, "bold"), text_color=COLORS.accent).grid(row=0, column=0, columnspan=2, pady=(20, 15))

        # Columna Izquierda: Identificación
        f_id = ctk.CTkFrame(card, fg_color="transparent")
        f_id.grid(row=1, column=0, sticky="nsew", padx=(40, 20), pady=(0, 25))
        items = [
            (f"🏢 Entidad: {self.app.session.get('nombre', '---')}",),
            (f"⚡ Fuente: {self.app.session.get('fuente', '---')}",),
            (f"📐 Unidad: {self.app.session.get('unidad', '---')}",),
            (f"🌍 Zona: {self.app.session.get('zona', '---')}",),
            (f"📏 Área Útil: {self.app.session.get('area', '---')}",),
            (f"📅 Periodo Base: {self.app.session.get('pb_ini', '---')} - {self.app.session.get('pb_fin', '---')}",)
        ]
        for txt, in items:
            ctk.CTkLabel(f_id, text=txt, font=(FONTS.family, FONTS.size_xs), text_color=COLORS.text_primary, anchor="w").pack(fill="x", pady=2)

        # Columna Derecha: Estadísticas de Carga
        f_st = ctk.CTkFrame(card, fg_color="transparent")
        f_st.grid(row=1, column=1, sticky="nsew", padx=(20, 40), pady=(0, 25))
        stats = [
            (f"📊 Registros Periodo Base: {self.app.session.get('n_pb', 0)} datos",),
            (f"📈 Registros Monitoreo: {self.app.session.get('n_pr', 0)} datos",),
            (f"📋 Columnas encontradas: {self.app.session.get('n_cols', 0)}",)
        ]
        for txt, in stats:
            ctk.CTkLabel(f_st, text=txt, font=(FONTS.family, FONTS.size_xs), text_color=COLORS.text_secondary, anchor="w").pack(fill="x", pady=2)

        # Botón Calcular Lime (Estilo M1)
        self.btn_calc = ctk.CTkButton(self.zona_resumen, text="🚀 Calcular Linea Base y Desempeño M2", font=(FONTS.family, FONTS.size_md, "bold"),
                                      fg_color=COLORS.accent, text_color=COLORS.primary, height=48, command=self._iniciar_procesamiento)
        self.btn_calc.grid(row=2, column=0, pady=30)

    def _iniciar_procesamiento(self):
        if self.df_base is None: return
        self.btn_calc.configure(state="disabled", text="PROCESANDO MODELO...")
        threading.Thread(target=self._hilo_procesar, daemon=True).start()

    def _hilo_procesar(self):
        try:
            df_lben, df_mon, df_bf, df_exc = procesar_m2(self.df_base, self.df_monitoreo)
            metricas = calcular_resumen_metricas_m2(df_lben, self.df_base)
            
            self.app.session.update({
                "df_base_raw": self.df_base,
                "df_base_final": df_bf,
                "df_excluidos": df_exc,
                "df_lben": df_lben,
                "df_monitoreo": df_mon,
                "metricas_m2": metricas,
                "excel_path": self.current_path
            })
            self.after(0, lambda: self.app.navegar("m2_resultados"))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Error", f"Fallo en el cálculo: {e}"))
            self.after(0, lambda: self.btn_calc.configure(state="normal", text="🚀 Reintentar Cálculo"))
