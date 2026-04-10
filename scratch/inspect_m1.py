import pandas as pd
import openpyxl

def inspect_m1_template(path):
    print(f"--- Inspeccionando Plantilla M1: {path} ---")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"Hojas encontradas: {wb.sheetnames}")
        
        for shelf in ['Periodo_Base', 'Modelo_LBEn', 'Monitoreo']:
            if shelf in wb.sheetnames:
                ws = wb[shelf]
                print(f"\nHoja: {shelf}")
                # Leer las primeras filas para confirmar encabezados
                for row in range(4, 9):
                    vals = [ws.cell(row=row, column=col).value for col in range(2, 14)]
                    print(f"Fila {row}: {vals}")
            else:
                print(f"\nAVISO: La hoja '{shelf}' no existe.")
                
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_m1_template("k:/gdrive_luisflorez/Mi unidad/PROYECTOS_LF/ProyectosE2/LBEn_EDIF_app/LBEn_APP_Resol016/contexto/consumo/Plantilla_LBEn_M1_modelo.xlsx")
