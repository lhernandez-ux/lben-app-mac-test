"""
core/io_excel.py
================
Lectura y escritura de archivos Excel para LBEn.
Soporta Análisis Exploratorio, Modelo M1, M2 y M3.
"""

import os
import shutil
from tkinter import filedialog, messagebox
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Rutas a plantillas ────────────────────────────────────────────────────────
_DIR_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DIR_DATA = os.path.join(_DIR_BASE, "data")
_PLANTILLA_M1 = os.path.join(_DIR_DATA, "Plantilla_LBEn_M1_modelo.xlsx")
_PLANTILLA_M2 = os.path.join(_DIR_DATA, "Plantilla_LBEn_M2_modelo.xlsx")
_PLANTILLA_M3 = os.path.join(_DIR_DATA, "Plantilla_LBEn_M3_modelo.xlsx")
_PLANTILLA_EXPLORATORIA = os.path.join(_DIR_DATA, "plantilla_exploracion_modelo.xlsx")

# ═══════════════════════════════════════════════════════════════════════════════
# A — GENERACIÓN DE PLANTILLAS
# ═══════════════════════════════════════════════════════════════════════════════

def generar_plantilla_exploratoria(nombre_proyecto, fecha_ini, fecha_fin, var_dep, vars_ind):
    if not os.path.exists(_PLANTILLA_EXPLORATORIA):
        messagebox.showerror("Error", f"No se encontró la plantilla.")
        return False
    nombre_archivo = f"exploracion_{_limpiar_nombre(nombre_proyecto)}.xlsx"
    ruta_destino = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=nombre_archivo)
    if not ruta_destino: return False
    shutil.copy2(_PLANTILLA_EXPLORATORIA, ruta_destino)
    try:
        wb = load_workbook(ruta_destino)
        _escribir_hoja_periodo(wb, fecha_ini, fecha_fin, var_dep, vars_ind)
        wb.save(ruta_destino)
        messagebox.showinfo("Éxito", "Plantilla generada.")
        return True
    except Exception as e:
        messagebox.showerror("Error", str(e))
        return False

def generar_plantilla_m1(data: dict) -> bool:
    return _generar_plantilla_generic(_PLANTILLA_M1, f"M1_{_limpiar_nombre(data['nombre'])}.xlsx", data, 1)

def generar_plantilla_m2(data: dict) -> bool:
    return _generar_plantilla_generic(_PLANTILLA_M2, f"M2_{_limpiar_nombre(data['nombre'])}.xlsx", data, 2)

def generar_plantilla_m3(data: dict) -> bool:
    return _generar_plantilla_generic(_PLANTILLA_M3, f"M3_{_limpiar_nombre(data['nombre'])}.xlsx", data, 3)

def _generar_plantilla_generic(path_tpl, default_name, data, mod_num):
    if not os.path.exists(path_tpl):
        messagebox.showerror("Error", "No se encontró la plantilla.")
        return False
    ruta_destino = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name)
    if not ruta_destino: return False
    try:
        shutil.copy2(path_tpl, ruta_destino)
        wb = load_workbook(ruta_destino)
        if mod_num == 1:
            _escribir_m1_identificacion(wb, data)
            _escribir_m1_periodo_base(wb, data)
            _escribir_m1_monitoreo(wb, data)
        elif mod_num == 2:
            _escribir_m2_identificacion(wb, data)
            _escribir_m1_periodo_base(wb, data)
            _escribir_m1_monitoreo(wb, data)
        elif mod_num == 3:
            _escribir_m3_identificacion(wb, data)
            _escribir_m3_hojas_datos(wb, data, "Período_Base")
            _escribir_m3_hojas_datos(wb, data, "Monitoreo")
        wb.save(ruta_destino)
        messagebox.showinfo("Éxito", "Plantilla generada.")
        return True
    except Exception as e:
        messagebox.showerror("Error", str(e))
        return False

# ── Helpers Escritura de Plantilla ──────────────────────────────────────────

def _escribir_m1_identificacion(wb, data):
    ws = wb["Modelo_LBEn"]
    ws["D5"] = data["nombre"]
    ws["D6"] = data["fuente"]
    ws["D7"] = data["unidad"]
    ws["D8"] = data.get("zona")
    ws["D9"] = data.get("area")

def _escribir_m1_periodo_base(wb, data):
    ws = wb["Período_Base"]
    fechas = _generar_fechas_mensuales(data["pb_ini"], data["pb_fin"])
    for i, f_str in enumerate(fechas):
        ws[f"B{8+i}"] = f_str

def _escribir_m2_identificacion(wb, data):
    _escribir_m1_identificacion(wb, data)
    ws = wb["Modelo_LBEn"]
    ws["D10"] = data.get("var_relevante_nom")
    ws["D11"] = data.get("var_relevante_uni")

def _escribir_m3_identificacion(wb, data):
    _escribir_m1_identificacion(wb, data)
    ws = wb["Modelo_LBEn"]
    vars_ind = data.get("vars_ind", [])
    for i in range(5):
        # B10:C10 combinadas → escribir en B (celda maestra)
        ws[f"B{10+i}"] = vars_ind[i] if i < len(vars_ind) else "—"

def _escribir_m3_hojas_datos(wb, data, sheet_name):
    ws = wb[sheet_name]
    vars_ind = data.get("vars_ind", [])
    for i in range(5):
        col_let = get_column_letter(5 + i)   # E=Var1, F=Var2, ..., I=Var5
        if i < len(vars_ind):
            ws[f"{col_let}6"].value = vars_ind[i]
        else:
            ws[f"{col_let}6"].value = "—"
            ws.column_dimensions[col_let].hidden = True
    f_ini = data["pb_ini"] if sheet_name == "Período_Base" else data["pr_ini"]
    f_fin = data["pb_fin"] if sheet_name == "Período_Base" else "12/2050"
    fechas = _generar_fechas_mensuales(f_ini, f_fin)
    for i, f_str in enumerate(fechas):
        ws[f"B{8+i}"] = f_str

def _escribir_m1_monitoreo(wb, data):
    ws = wb["Monitoreo"]
    fechas = _generar_fechas_mensuales(data["pr_ini"], data["pr_fin"])
    for i, f_str in enumerate(fechas):
        ws[f"B{8+i}"] = f_str

def _escribir_hoja_periodo(wb, f_ini, f_fin, var_dep, vars_ind):
    ws = wb["Periodo_Análisis"]
    ws["C6"] = var_dep
    for i, var in enumerate(vars_ind or []):
        col = get_column_letter(4 + i)
        ws[f"{col}6"] = var
    fechas = _generar_fechas_mensuales(f_ini, f_fin)
    for i, f_str in enumerate(fechas):
        ws[f"B{8+i}"] = f_str

def _generar_fechas_mensuales(f_ini, f_fin):
    meses_es = {1:"Ene", 2:"Feb", 3:"Mar", 4:"Abr", 5:"May", 6:"Jun",
                7:"Jul", 8:"Ago", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dic"}
    try:
        ini = datetime.strptime(f_ini, "%m/%Y")
        fin = datetime.strptime(f_fin, "%m/%Y")
        res = []
        curr = ini
        while curr <= fin:
            res.append(f"{meses_es[curr.month]}-{curr.year}")
            if curr.month == 12:
                curr = curr.replace(year=curr.year+1, month=1)
            else:
                curr = curr.replace(month=curr.month+1)
        return res
    except:
        return []

def _limpiar_nombre(n):
    import re
    return re.sub(r"[^\w\s-]", "", n.strip().lower()).replace(" ", "_")[:50]

def _clean_num(val):
    """Limpia strings numéricos con comas de miles y puntos decimales."""
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    try:
        # Ejemplo: "1,020.00" -> "1020.00"
        s = str(val).replace(",", "").strip()
        return float(s)
    except:
        return 0.0

# ═══════════════════════════════════════════════════════════════════════════════
# B — LECTURA DE DATOS
# ═══════════════════════════════════════════════════════════════════════════════

def leer_excel_exploratorio(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Periodo_Análisis"]
    headers = []
    c = 2
    while True:
        v = ws.cell(row=6, column=c).value
        if not v: break
        headers.append(str(v).strip()); c += 1
    datos = []
    r = 8
    while True:
        fv = ws.cell(row=r, column=2).value
        if not fv: break
        row = {"Fecha": str(fv)}
        for i, h in enumerate(headers[1:]):
            row[h] = ws.cell(row=r, column=3+i).value
        datos.append(row); r += 1
    df = pd.DataFrame(datos)
    return df, {"var_dep": headers[1], "vars_ind": headers[2:]}, []

def leer_excel_m1(path):
    wb = load_workbook(path, data_only=True)
    df_b = _leer_hoja_datos_generica(wb["Período_Base"])
    df_m = _leer_hoja_datos_generica(wb["Monitoreo"])
    ws_mod = wb["Modelo_LBEn"]
    meta = {
        "entidad": ws_mod["D5"].value,
        "fuente":  ws_mod["D6"].value,
        "unidad":  ws_mod["D7"].value,
        "zona":    ws_mod["D8"].value,
        "area":    ws_mod["D9"].value,
    }
    return df_b, df_m, meta, []

def leer_excel_m2(path):
    df_b, df_m, meta, err = leer_excel_m1(path)
    wb = load_workbook(path, data_only=True)
    ws_mod = wb["Modelo_LBEn"]
    meta.update({
        "var_relevante_nom": ws_mod["D10"].value,
        "var_relevante_uni": ws_mod["D11"].value
    })
    return df_b, df_m, meta, err

def _leer_hoja_datos_generica(ws):
    headers = ["Fecha"]
    for c in range(3, 15):
        v = ws.cell(row=6, column=c).value
        if v:
            headers.append(str(v).strip())
        else:
            break
    datos = []
    for r in range(8, 2000):
        fecha_v = ws.cell(row=r, column=2).value
        # Si no hay fecha, paramos inmediatamente
        if not fecha_v: break
        
        # Si hay fecha pero el consumo (columna 3) está vacío, paramos de leer datos reales
        # Esto evita cargar los 313 meses de la plantilla si solo hay 36 de data.
        consumo_v = ws.cell(row=r, column=3).value
        if consumo_v is None:
            break
            
        row = {"Fecha": str(fecha_v)}
        for i, h in enumerate(headers[1:]):
            val = ws.cell(row=r, column=3+i).value
            row[h] = _clean_num(val)
        datos.append(row)
    return pd.DataFrame(datos)

# ═══════════════════════════════════════════════════════════════════════════════
# C — ESCRITURA DE RESULTADOS
# ═══════════════════════════════════════════════════════════════════════════════

def escribir_resultados_exploratorios(path, rec, just, tabla):
    wb = load_workbook(path)
    ws_mod = wb["Modelo_LBEn"]
    ws_mod["C5"] = rec
    ws_mod["C7"] = just
    for i, row in enumerate(tabla):
        f = 13 + i
        ws_mod.cell(row=f, column=2, value=row.get("variable"))
        ws_mod.cell(row=f, column=3, value=row.get("r_pearson"))
        ws_mod.cell(row=f, column=4, value=row.get("p_valor"))
    wb.save(path)
    return True

def escribir_resultados_m1(path, df_lben, df_mon, df_b_f, df_excl, meta, config):
    wb = load_workbook(path)
    fmt = "#,##0.00"
    ws_mod = wb["Modelo_LBEn"]
    ws_mod["K5"] = "M1 (Consumo Absoluto)"
    ws_mod["M7"].value = meta.get("consumo_promedio_anual", 0)
    ws_mod["M7"].number_format = fmt
    ws_mod["M8"].value = meta.get("potencial_ahorro_kwh", 0)
    ws_mod["M8"].number_format = fmt
    ws_mod["M9"].value = meta.get("potencial_ahorro_pct", 0) / 100
    ws_mod["M9"].number_format = "0.0%"
    wb.save(path)
    return True

def escribir_resultados_m2(path, df_lben, df_mon, df_b_f, df_excl, meta, config):
    escribir_resultados_m1(path, df_lben, df_mon, df_b_f, df_excl, meta, config)
    wb = load_workbook(path)
    ws_mod = wb["Modelo_LBEn"]
    ws_mod["K5"] = "M2 (Cociente Normalizado)"
    wb.save(path)
    return True

def escribir_resultados_m3(path, res, config):
    """
    Escribe resultados del Modelo M3 en la plantilla Excel.
    Coordenadas validadas contra mapa real de la plantilla Plantilla_LBEn_M3_modelo.xlsx.
    """
    if not os.path.exists(path):
        return False
    wb = load_workbook(path)
    fmt_num = "#,##0.00"
    fmt_pct = "0.0%"

    from core.models.m3_regresion import formatear_ecuacion

    # ── 1. Hoja Período_Base — columnas calculadas Q, R, S ──────────────────
    ws_base = wb["Período_Base"]
    df_b = res['df_base'].reset_index(drop=True)
    for i, row in df_b.iterrows():
        f = 8 + i
        consumo = float(row.get('Consumo', 0) or 0)
        afect   = float(row.get('Afectacion', 0) or 0)
        pred    = float(row.get('y_pred_lben', 0) or 0)
        ajust   = consumo + afect
        desv    = ajust - pred
        ws_base[f"Q{f}"].value = ajust;  ws_base[f"Q{f}"].number_format = fmt_num
        ws_base[f"R{f}"].value = pred;   ws_base[f"R{f}"].number_format = fmt_num
        ws_base[f"S{f}"].value = desv;   ws_base[f"S{f}"].number_format = fmt_num

    # ── 2. Hoja Diag_Modelo ─────────────────────────────────────────────────
    ws_diag = wb["Diag_Modelo"]
    m = res['metrics']

    # 2a. Resumen estadístico global (col C, filas 7-11)
    estadisticos = {
        "C7":  m['r2'],
        "C8":  m['r2_adj'],
        "C9":  m['rmse'],
        "C10": m['cv_rmse'],
        "C11": m['f_pval'],
    }
    for cell, val in estadisticos.items():
        ws_diag[cell].value = val
        ws_diag[cell].number_format = fmt_num

    # 2b. Tabla de coeficientes (filas 15-20)
    # B15=Intercepto(β₀), B16=Variable1... son labels fijos → solo valores en C,D,E,F
    ct = res['coef_table']
    for i in range(len(ct['vars'])):
        f = 15 + i
        ws_diag[f"C{f}"].value = ct['betas'][i];    ws_diag[f"C{f}"].number_format = fmt_num
        ws_diag[f"D{f}"].value = ct['std_err'][i];  ws_diag[f"D{f}"].number_format = fmt_num
        ws_diag[f"E{f}"].value = ct['p_vals'][i];   ws_diag[f"E{f}"].number_format = fmt_num
        vif_val = ct['vif'][i]
        try:
            if vif_val is not None and not np.isnan(float(vif_val)):
                ws_diag[f"F{f}"].value = vif_val
                ws_diag[f"F{f}"].number_format = fmt_num
        except (TypeError, ValueError):
            pass

    # 2c. Ecuación LBEn en B21 (combinada B21:F21)
    ws_diag["B21"].value = formatear_ecuacion(res['model_lben'], config['vars_ind'])

    # 2d. Correlaciones de Pearson (filas 26-30)
    correls = res.get('correls', [])
    grado_labels = {0.9: "Fuerte", 0.7: "Moderado", 0.5: "Débil", 0.0: "Muy débil"}
    for i, cor in enumerate(correls):
        f = 26 + i
        r_val = abs(cor['r'])
        grado = next((v for k, v in sorted(grado_labels.items(), reverse=True) if r_val >= k), "Muy débil")
        ws_diag[f"C{f}"].value = cor['r'];  ws_diag[f"C{f}"].number_format = fmt_num
        ws_diag[f"D{f}"].value = cor['p'];  ws_diag[f"D{f}"].number_format = fmt_num
        ws_diag[f"E{f}"].value = grado

    # 2e. Breusch-Pagan (fila 36)
    ws_diag["D36"].value = m['bp_pval']
    ws_diag["D36"].number_format = fmt_num
    ws_diag["E36"].value = "Varianza constante" if m['bp_pval'] > 0.05 else "Heterocedasticidad detectada"

    # 2f. Tabla de datos excluidos estadísticamente (B30:D en adelante)
    # Plantilla: B27=título, B28=descripción, B29/D29=cabeceras, B30+=datos
    df_excl = res.get('df_excluidos', pd.DataFrame())
    if not df_excl.empty:
        df_excl = df_excl.reset_index(drop=True)
        for i, row in df_excl.iterrows():
            f = 30 + i
            ws_diag[f"B{f}"].value = str(row.get("Fecha", ""))
            ws_diag[f"D{f}"].value = float(pd.to_numeric(row.get("Consumo", 0), errors='coerce') or 0)
            ws_diag[f"D{f}"].number_format = fmt_num


    # ── 3. Hoja Modelo_LBEn ─────────────────────────────────────────────────
    # Mapa confirmado por inspección de plantilla:
    #   B:C combinadas = etiquetas | D = valor identificación
    #   I:J combinadas = etiquetas métricas | K = valor métrica
    #   L = etiqueta potencial | M = valor potencial
    ws_mod = wb["Modelo_LBEn"]
    p = res['potenciales']

    # Datos del proyecto (columna D)
    ws_mod["D5"] = config.get("nombre", "")
    ws_mod["D6"] = config.get("fuente", "")
    ws_mod["D7"] = config.get("unidad", "")
    ws_mod["D8"] = config.get("zona",   "")
    ws_mod["D9"] = config.get("area",   "No disponible")

    # Variables independientes (D10-D14)
    vars_ind = config.get("vars_ind", [])
    for i in range(5):
        ws_mod[f"D{10+i}"].value = vars_ind[i] if i < len(vars_ind) else ""

    # Métricas (columna K, filas 6-10)
    # K6: contar directamente en el Excel (col C = Consumo), no desde los DataFrames
    ws_pb_chk = wb["Período_Base"]
    n_total = 0
    fecha_ini_real = ""
    fecha_fin_real = ""
    for r in range(8, 2000):
        consumo_val = ws_pb_chk.cell(row=r, column=3).value   # Col C = Consumo
        fecha_val   = ws_pb_chk.cell(row=r, column=2).value   # Col B = Fecha
        if consumo_val is None and fecha_val is None:
            break
        if consumo_val is not None:
            n_total += 1
            if not fecha_ini_real and fecha_val:
                fecha_ini_real = str(fecha_val)
            if fecha_val:
                fecha_fin_real = str(fecha_val)

    df_excl  = res.get('df_excluidos', pd.DataFrame())
    n_excl   = len(df_excl)
    n_usados = len(df_b)
    fiabilidad = (n_usados / n_total * 100) if n_total > 0 else 0  # K9/K6 × 100

    ws_mod["K6"].value = n_total
    ws_mod["K7"].value = n_excl
    ws_mod["K8"].value = 0          # filtrado manual (no implementado)
    ws_mod["K9"].value = n_usados
    ws_mod["K10"].value = round(fiabilidad, 2)
    ws_mod["K10"].number_format = fmt_num

    ws_mod["M5"].value  = fecha_ini_real                    # Período base (inicio)
    ws_mod["M6"].value  = fecha_fin_real                    # Período base (fin)
    ws_mod["M7"].value  = p['prom_real'] * 12               # Consumo promedio anual (kWh)
    ws_mod["M7"].number_format  = fmt_num
    ws_mod["M8"].value  = p['ahorro_kwh'] * 12              # Potencial ahorro anual (kWh)
    ws_mod["M8"].number_format  = fmt_num
    ws_mod["M9"].value  = p['ahorro_pct'] / 100             # Potencial ahorro (%)
    ws_mod["M9"].number_format  = fmt_pct
    ws_mod["M10"].value = p['prom_real'] * 12 * 0.15        # Meta 15% anual (kWh) — Ley 2294/2023
    ws_mod["M10"].number_format = fmt_num

    # Tabla LB y Ahorro Potencial (columna D, filas 17-23)
    ws_mod["D17"].value = formatear_ecuacion(res['model_lben'], config['vars_ind'])
    ws_mod["D18"].value = formatear_ecuacion(res['model_lmen'], config['vars_ind'])
    ws_mod["D19"].value = p['prom_real'];    ws_mod["D19"].number_format = fmt_num
    ws_mod["D20"].value = p['prom_lben'];   ws_mod["D20"].number_format = fmt_num
    ws_mod["D21"].value = p['prom_lmen'];   ws_mod["D21"].number_format = fmt_num
    ws_mod["D22"].value = p['ahorro_kwh'];  ws_mod["D22"].number_format = fmt_num
    ws_mod["D23"].value = p['ahorro_pct'] / 100; ws_mod["D23"].number_format = fmt_pct

    # ── 4. Hoja Monitoreo — columnas calculadas R a AB ──────────────────────
    # REGLA: solo calcular filas donde hay consumo en col C.
    # LBEn (S) viene de df_mon calculado por el modelo.
    # Afectación (M), Tarifa (J), Factor Emisión (K) se leen del workbook (datos del usuario).
    ws_mon = wb["Monitoreo"]
    df_mon = res.get('df_mon', pd.DataFrame())

    pot_anual  = p['ahorro_kwh'] * 12
    meta_anual = p['prom_real'] * 12 * 0.15

    def _extraer_año_local(fecha_val):
        try:
            s = str(fecha_val).replace("/", "-")
            for parte in s.split("-"):
                if len(parte) == 4 and parte.isdigit():
                    return int(parte)
        except Exception:
            pass
        return None

    cusum_acum = 0.0
    año_actual = None
    df_mon_reset = df_mon.reset_index(drop=True) if not df_mon.empty else pd.DataFrame()

    for i in range(0, max(len(df_mon_reset), 2000)):
        fila_excel = 8 + i

        # Leer consumo de col C directamente del workbook (usuario lo escribe)
        consumo_raw = ws_mon.cell(row=fila_excel, column=3).value  # C = Consumo_kWh
        if consumo_raw is None:
            break  # No más datos → stop

        consumo_r = _clean_num(consumo_raw)
        if consumo_r <= 0:
            continue  # Fila sin consumo real → saltar

        # Leer datos del usuario directamente del workbook
        fecha_raw  = ws_mon.cell(row=fila_excel, column=2).value   # B = Fecha
        afect_raw  = ws_mon.cell(row=fila_excel, column=13).value  # M = Afectación NR
        tarifa_raw = ws_mon.cell(row=fila_excel, column=10).value  # J = Tarifa
        factor_raw = ws_mon.cell(row=fila_excel, column=11).value  # K = Factor Emisión

        afect_r  = _clean_num(afect_raw)
        tarifa_r = _clean_num(tarifa_raw)
        factor_r = _clean_num(factor_raw)

        # LBEn viene del df_mon calculado por el motor de regresión
        lben_r = 0.0
        if not df_mon_reset.empty and i < len(df_mon_reset):
            lben_raw = df_mon_reset.iloc[i].get('lben_mes', 0)
            lben_r = float(pd.to_numeric(lben_raw, errors='coerce') or 0)

        # CUSUM anual: reiniciar al cambiar de año
        año_fila = _extraer_año_local(fecha_raw)
        if año_fila is not None and año_fila != año_actual:
            cusum_acum = 0.0
            año_actual = año_fila

        # Cálculos según especificación:
        R = consumo_r + afect_r              # R = C + M
        S = lben_r                           # S = LBEn mes (modelo)
        T = R - S                            # T = R - S (negativo=ahorro, positivo=sobreconsumo)
        U = (T / S * 100) if S > 0 else 0   # U = T/S × 100 (%)
        cusum_acum += T                      # V acumula T, reinicia cada año
        V = cusum_acum
        W = (V / pot_anual  * 100 * -1) if pot_anual  != 0 else 0  # W = V/M8×100×-1
        X = (V / meta_anual * 100 * -1) if meta_anual != 0 else 0  # X = V/M10×100×-1
        Y  = T * tarifa_r                    # Y = T × J
        Z  = V * tarifa_r                    # Z = V × J
        AA = T * factor_r                    # AA = T × K
        AB = V * factor_r                    # AB = V × K

        ws_mon.cell(row=fila_excel, column=18).value         = R   # R
        ws_mon.cell(row=fila_excel, column=18).number_format = fmt_num
        ws_mon.cell(row=fila_excel, column=19).value         = S   # S
        ws_mon.cell(row=fila_excel, column=19).number_format = fmt_num
        ws_mon.cell(row=fila_excel, column=20).value         = T   # T
        ws_mon.cell(row=fila_excel, column=20).number_format = fmt_num
        ws_mon.cell(row=fila_excel, column=21).value         = U / 100  # U (%)
        ws_mon.cell(row=fila_excel, column=21).number_format = fmt_pct
        ws_mon.cell(row=fila_excel, column=22).value         = V   # V
        ws_mon.cell(row=fila_excel, column=22).number_format = fmt_num
        ws_mon.cell(row=fila_excel, column=23).value         = W / 100  # W (%)
        ws_mon.cell(row=fila_excel, column=23).number_format = fmt_pct
        ws_mon.cell(row=fila_excel, column=24).value         = X / 100  # X (%)
        ws_mon.cell(row=fila_excel, column=24).number_format = fmt_pct
        ws_mon.cell(row=fila_excel, column=25).value         = Y   # Y
        ws_mon.cell(row=fila_excel, column=25).number_format = fmt_num
        ws_mon.cell(row=fila_excel, column=26).value         = Z   # Z
        ws_mon.cell(row=fila_excel, column=26).number_format = fmt_num
        ws_mon.cell(row=fila_excel, column=27).value         = AA  # AA
        ws_mon.cell(row=fila_excel, column=27).number_format = fmt_num
        ws_mon.cell(row=fila_excel, column=28).value         = AB  # AB
        ws_mon.cell(row=fila_excel, column=28).number_format = fmt_num

    try:
        wb.save(path)
        return True
    except Exception as e:
        print(f"Error guardando Excel M3: {e}")
        return False


